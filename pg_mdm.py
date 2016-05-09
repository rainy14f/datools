# -*- coding: utf-8 -*-


from openpyxl import load_workbook

MASTER_DATA = []

KEY_COLUMNS = {
    'ref-key': 3,  # BAR_CODE
    'key': 1  # UPC
}

UPDATE_COLUMNS = [
    {'ref-col': 1, 'col': 9},  # PRODUCT_NAME_CN
    {'ref-col': 10, 'col': 11},  # BRAND
    {'ref-col': 6, 'col': 12},  # CATEGORY
    {'ref-col': 8, 'col': 13},  # SUB-CATEGORY
]


def load_master_data(file_path, sheet_name):
    workbook = load_workbook(filename=file_path, read_only=True)
    worksheet = workbook.get_sheet_by_name(sheet_name)
    row_index = 0
    for row in worksheet.rows:
        row_index += 1
        MASTER_DATA.append([v(c) for c in row])


def v(cell):
    if cell and cell.value:
        return unicode(cell.value)
    else:
        return ''


def vv(row, index):
    return v(row[index])


def get_row(ref_key, key_value):
    rows = [m for m in MASTER_DATA if compare_keys(m[ref_key], key_value) and key_value and m[ref_key]]
    return rows[0] if rows else None


def compare_keys(a, b):
    return a.find(b)>=0 or b.find(a)>=0


from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl import Workbook


def update():
    load_master_data('I:\\CID\\From customer\\201603 Local product hierarchy.xlsx', 'SQL Results')
    workbook = load_workbook(filename='I:\\CID\\Customer Inputs Document (CID) - CN (6)_updated 0509.xlsx')
    sheet = workbook.get_sheet_by_name(u'Tmall_FS本品')
    out_book = Workbook()
    out_sheet = out_book.create_sheet()
    header = True
    for row in sheet.rows:
        out_sheet.append([v(cell) for cell in row])
        out_row = out_sheet.rows[-1]
        if header:
            header = False
            continue
        else:
            key_value = vv(row, KEY_COLUMNS['key'])
            ref_row = get_row(KEY_COLUMNS['ref-key'], key_value)
            if key_value:
                print key_value
            if ref_row:
                for upd in UPDATE_COLUMNS:
                    col = upd['col']
                    ref_col = upd['ref-col']
                    if out_row[col].value != ref_row[ref_col]:
                        out_row[col].value = ref_row[ref_col]
                        out_row[col].font = Font(color=RED)

    out_book.save('I:\\out.xlsx')


update()

